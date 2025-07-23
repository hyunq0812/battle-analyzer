import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from PIL import Image
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Alignment # Import Alignment for cell formatting
import cv2 # OpenCV for image preprocessing
import numpy as np # For converting PIL Image to OpenCV format
import easyocr # Import EasyOCR

# --- Tesseract Path and TESSDATA_PREFIX Configuration (Kept for fallback, though EasyOCR is primary) ---
# 1. Specify the exact path to the Tesseract executable.
#    Please change the path according to your installation, e.g., 'C:\Program Files\Tesseract-OCR\tesseract.exe'.
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 2. Set the Tesseract 'tessdata' folder path in environment variables.
#    Typically, the tessdata folder is inside the folder where tesseract.exe is located.
#    Example: C:\Program Files\Tesseract-OCR\tessdata
os.environ['TESSDATA_PREFIX'] = r'C:\Program Files\Tesseract-OCR\tessdata'
# --- End Tesseract Configuration ---

# --- EasyOCR Reader Initialization (Global for efficiency) ---
# Initialize EasyOCR reader once when the script starts.
# 'en' for English digits. Set gpu=True if you have CUDA and want to use GPU for faster processing.
# For numbers, the 'en' model is usually sufficient and robust.
# If EasyOCR initialization fails (e.g., missing dependencies, no GPU), it will fall back to Tesseract.
easyocr_reader = None
try:
    # Setting gpu=False to use CPU for processing. This often resolves
    # issues related to CUDA/GPU driver/PyTorch version incompatibilities.
    easyocr_reader = easyocr.Reader(['en'], gpu=False) 
    print("EasyOCR reader initialized successfully (using CPU).")
except Exception as e:
    print(f"Warning: EasyOCR initialization failed ({e}). Falling back to Tesseract for OCR.")
    # EasyOCR will remain None, triggering Tesseract fallback in ocr_image.

def ocr_image(image_path):
    """
    Recognizes only numbers from an image and returns the text.
    Prioritizes EasyOCR for better accuracy. Falls back to Tesseract if EasyOCR fails or is not initialized.
    """
    try:
        # Read image directly with OpenCV, as EasyOCR and Tesseract (after conversion) can use it.
        img_cv = cv2.imread(image_path) 
        if img_cv is None:
            raise FileNotFoundError(f"Image file not found or corrupted: {image_path}")

        # --- Attempt with EasyOCR ---
        if easyocr_reader:
            try:
                # detail=1 returns bounding boxes, text, and confidence scores
                # allowlist='0123456789' restricts recognition to digits.
                results = easyocr_reader.readtext(img_cv, allowlist='0123456789') 
                
                # Iterate through all recognized results and return the first purely digit string found.
                # No confidence threshold applied, as per user request to revert.
                for (bbox, text, prob) in results:
                    if text.isdigit(): 
                        return text # Return the first recognized digit string
                
                # If EasyOCR found no purely digit results, fall through to Tesseract.
                print(f"EasyOCR found no valid digits for {image_path}. Attempting with Tesseract.")
            except Exception as e:
                print(f"EasyOCR processing error for {image_path}: {e}. Falling back to Tesseract.")
                # Continue to Tesseract fallback

        # --- Fallback to Tesseract (with previous preprocessing) ---
        # Convert OpenCV BGR to PIL Image for Tesseract
        # Apply CLAHE and Gaussian blur as these proved helpful with Tesseract for contrast.
        gray_img = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced_gray_img = clahe.apply(gray_img)
        blurred_img = cv2.GaussianBlur(enhanced_gray_img, (3, 3), 0)
        processed_img_pil = Image.fromarray(blurred_img)

        custom_config = r'-c tessedit_char_whitelist=0123456789 --psm 7 --oem 3' 
        text = pytesseract.image_to_string(processed_img_pil, config=custom_config)
        return text.strip()

    except Exception as e:
        print(f"Critical Error during OCR for {image_path}: {e}")
        return "" # Return empty string on critical error

def get_number_from_ocr_text(text):
    """
    Extracts the first sequence of digits from the OCR text and converts it to an integer.
    Returns 0 if no digits are found.
    Assumes each ROI image contains primarily one number.
    """
    numbers = re.findall(r'\d+', text)
    if numbers:
        return int(numbers[0])
    return 0

class OCRProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("OCR Analysis and Excel Generator")

        self.trillion_folder = tk.StringVar(value="")
        self.hundred_million_folder = tk.StringVar(value="")
        self.start_frame_num = tk.StringVar(value="0")
        self.frame_interval = tk.StringVar(value="24") # Default to 24 frames
        self.video_fps = tk.StringVar(value="23.84") # Default FPS

        self.create_widgets()

    def create_widgets(self):
        # Input Folders Frame
        folder_frame = tk.LabelFrame(self.root, text="Input Image Folders")
        folder_frame.pack(pady=10, padx=10, fill="x")

        tk.Label(folder_frame, text="Trillion Unit Images Folder:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        tk.Label(folder_frame, textvariable=self.trillion_folder, wraplength=300, anchor="w").grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        tk.Button(folder_frame, text="Browse", command=lambda: self.select_folder(self.trillion_folder)).grid(row=0, column=2, padx=5, pady=2)

        tk.Label(folder_frame, text="Hundred Million Unit Images Folder:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        tk.Label(folder_frame, textvariable=self.hundred_million_folder, wraplength=300, anchor="w").grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        tk.Button(folder_frame, text="Browse", command=lambda: self.select_folder(self.hundred_million_folder)).grid(row=1, column=2, padx=5, pady=2)

        # Processing Settings Frame
        settings_frame = tk.LabelFrame(self.root, text="Processing Settings")
        settings_frame.pack(pady=10, padx=10, fill="x")

        tk.Label(settings_frame, text="Start Frame Number:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        tk.Entry(settings_frame, textvariable=self.start_frame_num).grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        tk.Label(settings_frame, text="Frame Processing Interval:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        tk.Entry(settings_frame, textvariable=self.frame_interval).grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        tk.Label(settings_frame, text="Video FPS:").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        tk.Entry(settings_frame, textvariable=self.video_fps).grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        # Process Button
        tk.Button(self.root, text="Start OCR Analysis and Generate Excel", command=self.process_ocr).pack(pady=20)

        self.root.mainloop()

    def select_folder(self, var_string_obj):
        """Opens a directory dialog and sets the selected path to the StringVar."""
        path = filedialog.askdirectory(title="Select Folder")
        if path:
            var_string_obj.set(path)

    def process_ocr(self):
        """
        Main function to orchestrate OCR processing and Excel generation.
        """
        trillion_dir = self.trillion_folder.get()
        hundred_million_dir = self.hundred_million_folder.get()
        
        if not os.path.isdir(trillion_dir) or not os.path.isdir(hundred_million_dir):
            messagebox.showerror("Input Error", "Please select valid image folders for both Trillion and Hundred Million units.")
            return

        try:
            start_frame = int(self.start_frame_num.get())
            frame_interval = int(self.frame_interval.get())
            fps = float(self.video_fps.get())
            if start_frame < 0 or frame_interval <= 0 or fps <= 0:
                messagebox.showerror("Input Error", "Start Frame Number must be non-negative, Interval and FPS must be positive.")
                return
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numbers for Start Frame, Interval, and FPS.")
            return

        # Get all image files and sort them by frame number
        trillion_files = sorted([f for f in os.listdir(trillion_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))], 
                                key=lambda x: int(re.search(r'frame_(\d+)', x).group(1)) if re.search(r'frame_(\d+)', x) else 0)
        hundred_million_files = sorted([f for f in os.listdir(hundred_million_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))],
                                       key=lambda x: int(re.search(r'frame_(\d+)', x).group(1)) if re.search(r'frame_(\d+)', x) else 0)

        # Create a mapping from original frame number to file path for quick lookup
        trillion_map = {int(re.search(r'frame_(\d+)', f).group(1)): os.path.join(trillion_dir, f) for f in trillion_files if re.search(r'frame_(\d+)', f)}
        hundred_million_map = {int(re.search(r'frame_(\d+)', f).group(1)): os.path.join(hundred_million_dir, f) for f in hundred_million_files if re.search(r'frame_(\d+)', f)}

        all_frame_numbers = sorted(list(set(trillion_map.keys()).union(set(hundred_million_map.keys()))))
        if not all_frame_numbers:
            messagebox.showwarning("No Images", "No image files found in the specified folders.")
            return

        # Determine the maximum frame number to process
        max_frame_num = max(all_frame_numbers)

        wb = Workbook()
        ws = wb.active
        ws.title = "Accumulated Damage Analysis"
        
        # Excel Header
        ws.append(["Frame Number", "Trillion", "Hundred Million", "Total Accumulated Damage (Trillion Units)", "Damage Change (Trillion Units per interval)"])

        # Set header row alignment to center
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Variables for damage tracking across resets
        overall_accumulated_damage = 0.0 # The final total damage including all resets
        # last_successful_ocr_value_in_current_phase: The last raw OCR value seen in the current phase (e.g., 1.0, then 2.0)
        # This is used to detect significant drops (resets). Initialized to 0.0.
        last_successful_ocr_value_in_current_phase = 0.0 
        
        # reset_offset: Accumulates the total value of damage before a reset, to be added to subsequent readings.
        # This maintains the overall accumulated damage across boss phase resets. Initialized to 0.0.
        reset_offset = 0.0 

        # last_valid_jo, last_valid_uk: Store the components (Trillion, Hundred Million) of the last successfully
        # logged overall damage. Used when OCR fails or value decreases unexpectedly (monotonicity).
        last_valid_jo = 0
        last_valid_uk = 0
        
        # previous_total_damage_for_diff: Stores the total accumulated damage from the previous *logged* row.
        # Used to calculate the 'Damage Change' for the current interval. Initialized to 0.0.
        previous_total_damage_for_diff = 0.0 

        # Define thresholds for reset detection
        # A true reset is detected if:
        # 1. We had a previous successful OCR value (not 0).
        # 2. The raw current OCR value is significantly smaller than the previous successful value (e.g., < 1% of previous).
        # 3. AND the raw current OCR value is also below a very small absolute threshold (e.g., close to zero).
        # This helps distinguish true resets (damage goes back to a small number, usually near 0) from OCR errors (e.g., missing a digit).
        RESET_DROP_FACTOR = 0.01 # e.g., current value is less than 1% of previous successful value
        SMALL_ABSOLUTE_RESET_THRESHOLD = 0.00001 # e.g., if damage drops to below 0.00001 Trillion (10 million), it's a candidate for reset
                                        # This is 10000 / 10000 = 1, so 100 million. 0.00001 is 10k.

        # Iterate through frames based on start_frame and frame_interval
        current_frame_to_process = start_frame
        while current_frame_to_process <= max_frame_num:
            trillion_val = 0
            hundred_million_val = 0
            
            # These will be the values actually logged to Excel for Trillion and Hundred Million columns
            trillion_val_for_excel = 0
            hundred_million_val_for_excel = 0

            current_total_damage_to_log = 0.0 # The final calculated total damage for this row
            damage_change = 0.0
            
            # Check if images for the current frame exist
            trillion_img_path = trillion_map.get(current_frame_to_process)
            hundred_million_img_path = hundred_million_map.get(current_frame_to_process)

            # Flag to track if a reset was detected in the current iteration
            is_reset_detected = False 

            if trillion_img_path and hundred_million_img_path:
                # Perform OCR on both images
                trillion_text = ocr_image(trillion_img_path)
                hundred_million_text = ocr_image(hundred_million_img_path)

                trillion_val = get_number_from_ocr_text(trillion_text)
                hundred_million_val = get_number_from_ocr_text(hundred_million_text)
                
                # Calculate the raw damage from current OCR readings (without reset offset yet)
                raw_current_damage_from_ocr = trillion_val + (hundred_million_val / 10000.0)

                # --- Debugging Print Statements ---
                print(f"\n--- Frame {current_frame_to_process} ---")
                print(f"  Raw OCR Trillion Text: '{trillion_text}', Value: {trillion_val}")
                print(f"  Raw OCR Hundred Million Text: '{hundred_million_text}', Value: {hundred_million_val}")
                print(f"  Raw Combined Damage from OCR: {raw_current_damage_from_ocr:.4f} Trillion")
                print(f"  Last Successful OCR Value in Current Phase: {last_successful_ocr_value_in_current_phase:.4f} Trillion")
                print(f"  Current Overall Accumulated Damage (before this frame): {overall_accumulated_damage:.4f} Trillion")
                print(f"  Current Reset Offset: {reset_offset:.4f} Trillion")
                # --- End Debugging Print Statements ---


                # Refined Reset detection logic:
                # A reset is detected if:
                # 1. We had a previous successful OCR value (not 0).
                # 2. The raw current OCR value is significantly smaller than the previous successful value (e.g., < 1% of previous).
                # 3. AND the raw current OCR value is also below a very small absolute threshold (e.g., close to zero).
                # This helps distinguish true resets (damage goes back to a small number, usually near 0) from OCR errors (e.g., missing a digit).
                if last_successful_ocr_value_in_current_phase > 0 and \
                   raw_current_damage_from_ocr < last_successful_ocr_value_in_current_phase * RESET_DROP_FACTOR and \
                   raw_current_damage_from_ocr < SMALL_ABSOLUTE_RESET_THRESHOLD:
                    
                    is_reset_detected = True
                    # A reset is detected. The 'overall_accumulated_damage' up to the previous frame
                    # becomes the new 'reset_offset'.
                    reset_offset = overall_accumulated_damage
                    print(f"  --> RESET DETECTED! New reset_offset set to {reset_offset:.4f} Trillion.")
                    # Reset last_successful_ocr_value_in_current_phase for the new phase
                    last_successful_ocr_value_in_current_phase = 0.0 # It will be updated by the current raw_current_damage_from_ocr if valid
                
                # Calculate the current total damage for this frame, applying the reset offset.
                current_calculated_overall_damage = raw_current_damage_from_ocr + reset_offset

                # Now, apply the "accumulated damage never decreases" rule to the *overall* damage.
                # This handles minor OCR fluctuations that don't trigger a full reset,
                # or ensures monotonicity after a reset.
                if current_calculated_overall_damage < overall_accumulated_damage:
                    # If the calculated value is less than the *previously logged overall total*,
                    # it means this is an OCR error within the current phase, or a very slight dip.
                    # We use the previous overall_accumulated_damage to maintain monotonicity.
                    current_total_damage_to_log = overall_accumulated_damage
                    print(f"  --> WARNING: Calculated total ({current_calculated_overall_damage:.4f}) is less than previous overall ({overall_accumulated_damage:.4f}). Maintaining previous overall value.")
                    
                    # Use components from the last successfully logged overall damage for consistency in Excel
                    trillion_val_for_excel = last_valid_jo
                    hundred_million_val_for_excel = last_valid_uk
                else:
                    # If the calculated value is increasing or equal, update the overall accumulated damage.
                    overall_accumulated_damage = current_calculated_overall_damage
                    current_total_damage_to_log = overall_accumulated_damage # This is the value to log
                    print(f"  --> UPDATED Overall Accumulated Damage to: {overall_accumulated_damage:.4f} Trillion")

                    # Update last_successful_ocr_value_in_current_phase for the next iteration's reset detection.
                    # This should only be updated if the current raw OCR value is *meaningful* for phase tracking.
                    # If it was a reset, the raw_current_damage_from_ocr is the new baseline for the phase.
                    # If it was just an increase within a phase, update it.
                    # Also, only update if raw_current_damage_from_ocr is not 0, to avoid setting baseline to 0.
                    if raw_current_damage_from_ocr > 0 and (raw_current_damage_from_ocr >= last_successful_ocr_value_in_current_phase or is_reset_detected):
                        last_successful_ocr_value_in_current_phase = raw_current_damage_from_ocr
                        print(f"  Updated Last Successful OCR Value in Current Phase to: {last_successful_ocr_value_in_current_phase:.4f} Trillion")
                    
                    # Use current OCR components for display in Excel
                    trillion_val_for_excel = trillion_val
                    hundred_million_val_for_excel = hundred_million_val
                
                # Update last_valid_jo and last_valid_uk to reflect the components that led to the *logged* overall_accumulated_damage
                last_valid_jo = trillion_val_for_excel
                last_valid_uk = hundred_million_val_for_excel

            else:
                # If images for this frame are missing, use the last valid total damage
                print(f"\n--- Frame {current_frame_to_process} ---")
                print(f"  Info: Images for frame {current_frame_to_process} not found. Using last valid damage ({overall_accumulated_damage:.4f} Trillion).")
                current_total_damage_to_log = overall_accumulated_damage
                trillion_val_for_excel = last_valid_jo # Use the last valid component values
                hundred_million_val_for_excel = last_valid_uk


            # Calculate damage change
            # For the very first valid entry, damage_change is 0 or the initial value itself
            if previous_total_damage_for_diff == 0.0 and current_total_damage_to_log > 0.0:
                damage_change = current_total_damage_to_log # First valid damage is its own change from 0
            elif current_total_damage_to_log >= previous_total_damage_for_diff:
                damage_change = current_total_damage_to_log - previous_total_damage_for_diff
            else:
                # This case should ideally not happen if current_total_damage_to_log is always >= previous_total_damage_for_diff
                # due to the monotonicity safeguard. If it does, it's an edge case, so we set change to 0.0 to avoid negative.
                damage_change = 0.0
            
            previous_total_damage_for_diff = current_total_damage_to_log # Update for next iteration

            # Add data to Excel row
            row_data = [
                current_frame_to_process, 
                trillion_val_for_excel, 
                hundred_million_val_for_excel, 
                current_total_damage_to_log, 
                damage_change
            ]
            ws.append(row_data)
            
            # Apply center alignment to the newly added row
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            print(f"  Logged to Excel: Trillion={trillion_val_for_excel}, Hundred Million={hundred_million_val_for_excel}, Total={current_total_damage_to_log:.4f} Trillion, Change={damage_change:.4f} Trillion")

            current_frame_to_process += frame_interval

        # Define column widths after all data is appended
        # Adjust these widths as needed for your data
        column_widths = {
            'A': 15, # Frame Number
            'B': 12, # Trillion
            'C': 18, # Hundred Million
            'D': 28, # Total Accumulated Damage (Trillion Units)
            'E': 28  # Damage Change (Trillion Units per interval)
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save Excel file
        excel_save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                        filetypes=[("Excel files", "*.xlsx")],
                                                        title="Save Analysis Result Excel As")
        if excel_save_path:
            wb.save(excel_save_path)
            messagebox.showinfo("Complete", f"Analysis complete! Excel file saved to:\n{excel_save_path}")
        else:
            messagebox.showwarning("Cancelled", "Excel file save cancelled.")

if __name__ == "__main__":
    app = OCRProcessorApp()
